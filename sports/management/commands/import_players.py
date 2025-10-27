from openpyxl import load_workbook 
import os
from django.core.management.base import BaseCommand
from django.conf import settings
from mysite.settings import BASE_DIR
from sports.models import Team, Player 

class Command(BaseCommand):
    help = 'Import players from excel file'

    def handle(self, *args, **options):
        folderName = str(input("Enter folder name: "))

        xlsx_file_path = os.path.join(settings.BASE_DIR, 'data.xlsx')

        if not os.path.exists(xlsx_file_path):
            self.stdout.write(self.style.ERROR('XLSX file not found at: {}'.format(xlsx_file_path)))
            return

        workbook = load_workbook(xlsx_file_path)
        worksheet = workbook.active  # Assumes data is on the first sheet

        # Get headers from the first row
        headers = [cell.value for cell in worksheet[1]]
        self.stdout.write(self.style.SUCCESS(f'Headers found: {headers}'))

        # Iterate through rows (starting from row 2 to skip header)
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            # Create a dictionary mapping headers to row values
            row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}

            # --- Process each row ---
            team_name = row_dict.get('team')
            first_name = row_dict.get('first_name') or ''
            last_name = row_dict.get('last_name') or ''
            position = row_dict.get('position') or ''
            year_group_raw = row_dict.get('year_group')
            photo_filename = row_dict.get('photo') or ''
            is_captain_raw = row_dict.get('is_captain')
            kit_number_raw = row_dict.get('kit_number')
            quote = row_dict.get('Quote') or ''

            year_group = ''
            if year_group_raw is not None:
                year_group = str(year_group_raw).strip()

            # Validate required fields
            if not team_name or not first_name or not last_name:
                self.stdout.write(self.style.WARNING(f'Skipping row {row_num}: Missing required data (team, first_name, last_name). Values: {row_dict}'))
                continue

            # Get or create the Team
            try:
                team = Team.objects.get(name=team_name)
            except Team.DoesNotExist:
                self.stdout.write(self.style.WARNING(f'Row {row_num}: Team "{team_name}" not found. Skipping player: {first_name} {last_name}'))
                continue

            # Parse boolean for captain
            is_captain = False
            if is_captain_raw is not None:
                is_captain = str(is_captain_raw).strip().upper() in ['TRUE', '1', 'YES', 'Y']

            # Parse kit number (shirt_number)
            shirt_number = None
            if kit_number_raw is not None:
                try:
                    # Convert to int if it's a number-like value
                    shirt_number = int(kit_number_raw)
                except (ValueError, TypeError):
                    self.stdout.write(self.style.WARNING(f'Row {row_num}: Invalid kit number for {first_name} {last_name}: {kit_number_raw}'))

            # Create Player
            player, created = Player.objects.get_or_create(
                team=team,
                first_name=first_name.strip(),
                last_name=last_name.strip(),
                defaults={
                    'position': position.strip(),
                    'year': year_group.strip(),  # Note: model field is 'year', but your sheet says 'year_group'
                    'is_captain': is_captain,
                    'shirt_number': shirt_number,
                    'quote': quote.strip(),
                }
            )

            # Handle photo
            photo_filename = photo_filename.strip()
            if photo_filename:
                # Assuming photos are stored in 'media/players/photos/' or similar
                photo_path = os.path.join(settings.MEDIA_ROOT, 'players', 'photos', folderName, photo_filename)
                if os.path.exists(photo_path):
                    # Set the photo field
                    player.photo.name = f'players/photos/{folderName}/{photo_filename}'
                    print(player.photo.name)
                    player.save()
                else:
                    self.stdout.write(self.style.WARNING(f'Row {row_num}: Photo not found: {photo_filename} for {player}'))

            if created:
                self.stdout.write(self.style.SUCCESS(f'Row {row_num}: Created player: {player}'))
            else:
                self.stdout.write(self.style.WARNING(f'Row {row_num}: Player already exists: {player}'))

        self.stdout.write(self.style.SUCCESS('✅ Player import from XLSX completed.'))